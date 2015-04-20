import openpyxl
import time
from collections import deque
def howlong_script(f):
   def tmp(*args, **kwargs):
       t = time.time()
       res = f(*args, **kwargs)
       global time_
       time_  = (time.time()-t)
       return res
   return tmp

def howlong_all(f):
   def tmp(*args, **kwargs):
       t = time.time()
       res = f(*args, **kwargs)
       global time_all
       time_all  = (time.time()-t)
       return res
   return tmp

def adjacent_horse(matrix, x, y, finish):
    list_of_adjacent = []
    list_of_adjacent.append((x+2, y+1))
    list_of_adjacent.append((x+2, y-1))
    list_of_adjacent.append((x+1, y+2))
    list_of_adjacent.append((x-1, y+2))
    list_of_adjacent.append((x-2, y+1))
    list_of_adjacent.append((x-2, y-1))
    list_of_adjacent.append((x+1, y-2))
    list_of_adjacent.append((x-1, y-2))

    #Check if our figure can step on adjacent element:
    list_of_adjacent_final = []
    #Check if adjacent point is inside matrix
    for i in list_of_adjacent:
        if i[0] >=0 and i[1] >=0 and \
                i[0] <= finish[0] and i[1] <= finish[1]:
            #Check if horse can step there
            if matrix[i[0]][i[1]] == 0:
                list_of_adjacent_final.append(i)

    return list_of_adjacent_final

#Function to find the shortest way
@howlong_script
def bfs(matrix, start, finish):
    #Queue to check every point
    queue = []
    final_path = []
    queue.append(start)
    while (queue):
        # get the first path from the queue
        current_point = queue.pop(0)
        if current_point == finish:
            while current_point != start:
                #x = current_point
                final_path.append(current_point)
                current_point = matrix[current_point[0]][current_point[1]]
            return final_path[::-1]

        for i in (adjacent_horse(matrix, current_point[0], current_point[1], finish)):
            matrix[i[0]][i[1]] = current_point
            queue.append(i)

def find_start(ws):
    for row in ws.iter_rows():
        for i in row:
            if (i.border.left.style == 'medium') and (i.border.top.style == 'medium'):
                return i.column, i.row

def find_right_top(ws, row):
    for row in ws.iter_rows(row_offset=row-1):
        for i in row:
            if (i.border.right.style == 'medium') and (i.border.top.style == 'medium'):
                return i.column, i.row

def find_right_bottom(ws, row, column):
    range_string_= column + str(row) + ":" + column + str(ws.get_highest_row()+1)
    for row in ws.iter_rows(range_string = range_string_):
        for i in row:
            if (i.border.right.style == 'medium') and (i.border.bottom.style == 'medium'):
                return i.column, i.row



    # printing coordinates from "A3:E5" range

def create_matrix(ws, start, finish):
    matrix = []
    range_string_ = start[0] + str(start[1]) + ':' + finish[0] + str(finish[1])
    for row in ws.iter_rows(range_string = range_string_):
        matrix.append([])
        for i in row:
            if i.fill.fill_type == 'solid':
                matrix[i.row - start[1]].append(-1)
            else:
                 matrix[i.row - start[1]].append(0)
    return matrix

@howlong_all
def all():
    wb = openpyxl.load_workbook('kv008horse.xlsx')
    ws = wb.get_active_sheet()

    start = find_start(ws)
    right_top = find_right_top(ws, start[1])
    finish = find_right_bottom(ws, right_top[1], right_top[0])

    matrix = create_matrix(ws, start, finish)

    final = bfs(matrix, (0, 0), ((len(matrix))-1, len(matrix[0])-1))

    for i in final:
        ws.cell(row = i[0]+4, column = i[1]+2).value = "$"

    #Closing file + drawing path of the Horse
    for i in range(1, (openpyxl.cell.column_index_from_string(finish[0]))+1):
        d = ws.cell(row = 4, column = i)
        ws.column_dimensions[d.column].width = 2.1
    wb.save('kv008horse.xlsx')

all()

#Let's write time of work to our file
wb = openpyxl.load_workbook('kv008horse.xlsx')
ws = wb.get_active_sheet()
ws.cell('R2').value = round((time_ * 1000), 2)
ws.cell('AP2').value = round((time_ * 1000 + time_all * 1000),2)

#print (time_ * 1000)
#print (time_ * 1000 + time_all * 1000)
wb.save('kv008horse.xlsx')
